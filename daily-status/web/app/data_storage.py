# data_storage.py
import json
import math
from tkinter import messagebox
import openpyxl
from reportlab.lib.pagesizes import letter
from .utils import DATA_FILE, EXCEL_FILE, ADMIN_EXCEL_FILE
from datetime import date, datetime
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import calendar
from openpyxl.styles import Border, Side, PatternFill, Alignment, Font

# Global variables
students = {}
times = []
column1_options = {}
column2_options = {}
column3_options = {}
column4_options = {}
colors = {}
default_font = Font(name="Century Gothic", size=9)  # Font for all other cells

# Define border style
thin_border = Border(
    left=Side(style="thin", color="A9A9A9"),
    right=Side(style="thin", color="A9A9A9"),
    top=Side(style="thin", color="A9A9A9"),
    bottom=Side(style="thin", color="A9A9A9"),
)

# Map day numbers to single initials
day_initials = ["M", "T", "W", "R", "F", " ", " "]

def initialize_data():
    print("...in initialize_data")
    """Load data from the JSON file into global variables."""
    global students, times, column1_options, column2_options, column3_options, column4_options, colors
    try:
        with open(DATA_FILE, "r") as file:
            data = json.load(file)
            students = data.get("students", {})  # Load as a dictionary
            times = data.get("times", [])
            column1_options = data.get("column1_options", {})  # Load as a dictionary
            column2_options = data.get("column2_options", {})  # Load as a dictionary
            column3_options = data.get("column3_options", {})  # Load as a dictionary
            column4_options = data.get("column4_options", {})  # Load as a dictionary
            colors = data.get("colors", {})

    except FileNotFoundError:
        print(f"{DATA_FILE} not found. Using default values.")
        students = {}
        times = []
        column1_options = {}
        column2_options = {}
        column3_options = {}
        column4_options = {}
        colors = {}

def save_data():
    data = {
        "students": students,
        "times": times,
        "column1_options": column1_options,
        "column2_options": column2_options,
        "column3_options": column3_options,
        "column4_options": column4_options,
        "colors": colors,
    }
    update_data(data)

def update_data(data):
    try:
        with open(DATA_FILE, "w") as json_file:
            json.dump(data, json_file, indent=4)
        print(f"Data successfully saved to {DATA_FILE}")
    except Exception as e:
        print(f"Error saving data to {DATA_FILE}: {e}")
                   
def get_weekdays_and_weekends(year, month):
    # Get the total number of days in the month
    _, num_days = calendar.monthrange(year, month)

    weekdays = []  # List to store Monday-Friday dates
    weekends = []  # List 

    # Iterate through all days of the month
    for day in range(1, num_days + 1):
        day_of_week = date(year, month, day).weekday()  # 0 = Monday, 6 = Sunday
        day_initial = day_initials[day_of_week]  # Get the single initial for the day

        if day_of_week < 5:  # Monday-Friday
            weekdays.append((day, day_initial))
        else:  # Saturday-Sunday
            weekends.append((day, day_initial))

    return weekdays, weekends

def write_to_excel(data):
    global times
    print("...in write_to_excel")
    print(f"Data being written: {data}")
    # print(f"Looking for time {data['Time']} in sheet {sheet_name}")

    """Write data to an Excel file with student-specific sheets for each month."""
    # Get the current date
    today = datetime.now()
    current_month = today.strftime("%B")  # Full month name (e.g., "December")
    current_day = today.day  # Day of the month (e.g., 31)

    try:
        # Load the workbook if it exists, otherwise create a new one
        if os.path.exists(EXCEL_FILE):
            workbook = openpyxl.load_workbook(EXCEL_FILE)
        else:
            workbook = openpyxl.Workbook()
            # Remove the default sheet created by openpyxl
            default_sheet = workbook.active
            workbook.remove(default_sheet)

        # load admin version 
        # if os.path.exists(ADMIN_EXCEL_FILE):
        #     admin_workbook = openpyxl.load_workbook(ADMIN_EXCEL_FILE)
        # else:
        #     admin_workbook = openpyxl.Workbook()
        #     # Remove the default sheet created by openpyxl
        #     admin_default_sheet = admin_workbook.active
        #     admin_workbook.remove(admin_default_sheet)

        print("found workbooks")

        # Get the student's name
        student_name = data["Student"]

        # Create the sheet name for the current student and month
        sheet_name = f"{student_name}-{current_month}"

        print(f"Looking for time {data['Time']} in sheet {sheet_name}")

        # Check if the sheet already exists, otherwise create it
        if sheet_name not in workbook.sheetnames:
            print ("sheet not in workbook - create")
            sheet = workbook.create_sheet(title=sheet_name)
            # admin_sheet = admin_workbook.create_sheet(title=sheet_name)

            sheet = create_worksheet(sheet, student_name, current_month, today)
            # admin_sheet = create_worksheet(admin_sheet, student_name, current_month, today)
        else:
            # If the sheet already exists, load it
            print("sheet already exists")
            sheet = workbook[sheet_name]
            # admin_sheet = admin_workbook[sheet_name]

        # Find the column for the current day
        day_column = None
        for col in range(2, sheet.max_column + 1):  # Start from column B
            col_letter = openpyxl.utils.get_column_letter(col)
            cell_value = sheet[f"{col_letter}4"].value  # Get the value in row 4 for the current column
            print("cell_value: ", cell_value)

            # Ensure both values are integers for comparison
            if col is not None and int(cell_value) == int(current_day):
                print("found match day column")
                print("col: ", col)
                print("cell_Value: ", cell_value)
                print("current_day: ", current_day)
                day_column = col_letter
                break

        # Raise an error if the column is not found
        if day_column is None:
            raise ValueError(f"Could not find the column for day {current_day} in sheet {sheet.title}.")

        print("day_column:", day_column)

        # Find the row for the selected time
        time_row = None
        for row in range(5, sheet.max_row + 1):  # Start from row 5
            if sheet[f"A{row}"].value == data["Time"]:
                time_row = row
                break

        if not time_row:
            raise ValueError(f"Could not find the row for time {data['Time']} in sheet {sheet_name}.")

        print("time_row:", time_row)
        print(data)

        # Collect values from column1, column2, column3, column4, and notes
        values_to_concatenate = [
            data[key].strip() for key in ["column2", "column1", "column3", "column4"]
            if key in data and data[key] and data[key].strip() != "UNSELECTED"  # Exclude empty, None, or "UNSELECTED"
        ]
        print("values_to_concatenate:", values_to_concatenate)

        # Join the remaining values into a single string with a newline delimiter
        concatenated_values = "".join(values_to_concatenate)

        # Add "Notes" on a new line if it exists
        if "Notes" in data and data["Notes"] and data["Notes"].strip() != "UNSELECTED":
            concatenated_values += "\n" + data["Notes"].strip()  # Add Notes on a new line
        # admin_values = concatenated_values

        username = data.get("Username", "N/A")
        # Concatenate the username with the other values
        concatenated_values = "(" + username + ") " + concatenated_values
        print("concatenated values: ", concatenated_values)
        # print("admin values: ", admin_values)

        # Write the concatenated values to the appropriate cell
        sheet[f"{day_column}{time_row}"] = concatenated_values
        # admin_sheet[f"{day_column}{time_row}"] = admin_values
        sheet = format_worksheet(sheet, year=today.year, month=today.month)
        # admin_sheet = format_worksheet(admin_sheet, year=today.year, month=today.month)
        print("sheet: ", sheet)
        # print("admin sheet: ", admin_sheet)

        # Save the workbook
        workbook.save(EXCEL_FILE)
        # admin_workbook.save(ADMIN_EXCEL_FILE)
        print("done with excel")

        return True
    except Exception as e:
        print(f"Error updating {EXCEL_FILE}: {e}")

def create_worksheet(sheet, student_name, current_month, today):
    # Add the student's name in the first row
    sheet["A1"] = student_name
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    sheet["A2"] = current_month
    sheet["A2"].alignment = Alignment(horizontal="left", vertical="center")
    sheet["A3"] = today.year
    sheet["A3"].alignment = Alignment(horizontal="left", vertical="center")

    sheet["A1"].font = default_font
    sheet["A1"].alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    sheet["A1"].border = thin_border
    sheet["A2"].font = default_font
    sheet["A2"].alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    sheet["A2"].border = thin_border
    sheet["A3"].font = default_font
    sheet["A3"].alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    sheet["A3"].border = thin_border

    # Add headers in the fourth row
    sheet["A4"] = "Dates"
    sheet["A4"].font = default_font
    sheet["A4"].alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    sheet["A4"].border = thin_border

    # for day in range(1, 32):  # Maximum days in a month
    for day in range(1, calendar.monthrange(today.year, today.month)[1] + 1):
        col_letter = openpyxl.utils.get_column_letter(day + 1)  # Start from column B
        sheet[f"{col_letter}4"] = str(day)
        sheet[f"{col_letter}4"].font = default_font
        sheet[f"{col_letter}4"].alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        sheet[f"{col_letter}4"].border = thin_border

    print("Times list:", times)
    # Add times in column A starting from row 5
    for idx, time in enumerate(times, start=5):
        sheet[f"A{idx}"] = time
        sheet[f"A{idx}"].font = default_font
        sheet[f"A{idx}"].alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        sheet[f"A{idx}"].border = thin_border
    return sheet

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

def format_worksheet(sheet, year, month):
    """
    Apply formatting to the worksheet:
    - Enable text wrapping for all cells.
    - Shade weekend columns.
    - Set font to "Century Gothic", size 10.
    - Adjust column widths dynamically, ensuring no column is smaller than the width of the day of the month in row 3.
    """
    print("format_worksheet")
    # shade_weekends
    # Define fill colors for shading
    weekend_fill = PatternFill(start_color="dbdbdb", end_color="dbdbdb", fill_type="solid")  # Dark gray for weekends
    weekday_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White for weekdays

    # Get weekdays and weekends for the given month and year
    _, num_days = calendar.monthrange(year, month)

    # Write headers for each day of the month
    for day in range(1, num_days + 1):
        col_idx = day + 1  # Assuming column 1 is reserved for row labels
        col_letter = get_column_letter(col_idx)  # Get the column letter (e.g., B, C, etc.)
        day_of_week = date(year, month, day).weekday()  # 0 = Monday, 6 = Sunday

        max_length = 0  # Initialize the maximum length for the column
        has_data = False  # Flag to check if the column has any data

        # Write the letter day of the week in row 2
        sheet.cell(row=3, column=col_idx, value=day_initials[day_of_week])
        sheet.cell(row=3, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")
        sheet.cell(row=3, column=col_idx).font = default_font

        # Write the number day of the month in row 3
        sheet.cell(row=4, column=col_idx, value=day)
        sheet.cell(row=4, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")
        sheet.cell(row=4, column=col_idx).font = default_font

        # Set column width based on whether the day is a weekend or weekday
        if day_of_week >= 5:  # Saturday or Sunday
            sheet.column_dimensions[col_letter].width = 2.3  # Weekend column width (24 pixels)
        else:
            sheet.column_dimensions[col_letter].width = 4  # Weekday column width (75 pixels)

        # Apply shading and borders to all rows in the column
        for row_idx in range(4, sheet.max_row + 1):  # Start from row 2
            cell = sheet.cell(row=row_idx, column=col_idx)

            if day_of_week >= 5:  # Saturday or Sunday
                cell.fill = weekend_fill  # Apply weekend fill color
            else:
                cell.fill = weekday_fill  # Apply weekday fill color
            cell.font = default_font
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
            cell.border = thin_border

    for col in range(1, sheet.max_column + 1):  # Iterate through all columns
        col_letter = get_column_letter(col)  # Get the column letter (e.g., A, B, C)
        max_length = 0  # Initialize the maximum length for the column
        has_data = False  # Flag to check if the column has any data

        for row in range(5, sheet.max_row + 1):  # Iterate through all rows in the column
            cell = sheet.cell(row=row, column=col)
            if cell.value:  # Check if the cell has a value
                has_data = True  # Mark that the column has data
                # Split the cell value into lines based on wrapping (newlines)
                lines = str(cell.value).split("\n")
                # Find the longest line in the cell
                max_line_length = max(len(line) for line in lines)
                # Update the maximum length for the column
                max_length = max(max_length, max_line_length)

                sheet.row_dimensions[row].height = None

        # Only adjust the column width if the column has data
        if has_data:
            # Set the column width (add some padding for better readability)
            if max_length > 8:
                sheet.column_dimensions[col_letter].width = 8 

        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 1
        sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
        
        # Set margins to narrow
        sheet.page_margins.left = 0.25  # Narrow left margin
        sheet.page_margins.right = 0.25  # Narrow right margin
        sheet.page_margins.top = 0.25  # Narrow top margin
        sheet.page_margins.bottom = 0.25  # Narrow bottom margin
    return sheet