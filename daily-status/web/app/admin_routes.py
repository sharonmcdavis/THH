from datetime import datetime
import shutil
from flask import Blueprint, json, render_template, request, redirect, session, url_for, flash, send_file
import openpyxl
from .data_storage import save_data, update_data, format_worksheet
from .data_loader import load_data_from_file
from .utils import login_required, ADMIN_PASSWORD, EXCEL_FILE, ADMIN_EXCEL_FILE, ARCHIVE_FOLDER, admin_login_required
import os
import openpyxl
from flask import send_file
from openpyxl import load_workbook

admin = Blueprint('admin', __name__)

# Load data from the JSON file
data = load_data_from_file()

@admin.route('/')
@login_required
def admin_login():
    """Render the main page."""
    return render_template(
        'admin_login.html'
    )

@admin.route('/verify_admin', methods=['POST'])
@login_required
def verify_admin():
    # Get the password entered by the user
    entered_password = request.form.get('admin_password')

    # Verify the password
    if entered_password == ADMIN_PASSWORD:
        session['admin_logged_in'] = True
        return redirect(url_for('admin.admin_window'))  # Redirect to the admin page
    else:
        flash("Incorrect password. Access denied.", "error")
        return redirect(url_for('admin.admin_login'))  # Redirect back to the main page
    
@admin.route('/open')
@login_required
@admin_login_required
def admin_window():
    """Render the main page."""
    return load_admin_page()


@admin.route('/add_student', methods=['POST'])
@login_required
@admin_login_required
def add_student():
    """Add a new student."""
    if request.method == 'POST':
        from .data_storage import students
        student = request.form.get('student')
        color = request.form.get('color')  # Get the student color from the form

        if student and color:
            print("\nstudents before:", students)
            students[student] = color  # Assuming `students` is a dictionary
            save_data()
            print("saved student")
            flash(f"Student {student} added successfully!", "success")
        else:
            flash("Failed to add student. Please provide both name and color.", "error")

    print("\nstudents after:", students)

    return load_admin_page()


@admin.route('/remove_student', methods=['POST'])
@login_required
@admin_login_required
def remove_student():
    from .data_storage import students
    """Remove a student."""
    student = request.form.get('student')
    if student and student in students:
        del students[student]
        save_data()
        flash(f"Student '{student}' removed successfully!", "success")
    else:
        flash("Student not found.", "error")
    return load_admin_page()


@admin.route('/add_time', methods=['POST'])
@login_required
@admin_login_required
def add_time():
    from .data_storage import times
    """Add a new time."""
    time = request.form.get('time')
    if time and time not in times:
        times.append(time)
        save_data()
        flash(f"Time '{time}' added successfully!", "success")
    else:
        flash("Time already exists or is invalid.", "error")
    return load_admin_page()


@admin.route('/remove_time', methods=['POST'])
@login_required
@admin_login_required
def remove_time():
    from .data_storage import times
    """Remove a time."""
    time = request.form.get('time')
    if time and time in times:
        times.remove(time)
        save_data()
        flash(f"Time '{time}' removed successfully!", "success")
    else:
        flash("Time not found.", "error")
    
    return load_admin_page()


@admin.route('/add_column', methods=['POST'])
@login_required
@admin_login_required
def add_column():
    form_data = request.form.to_dict()
    print("form_data:", form_data)

    # Extract column name, key, and value
    column_name = form_data.get('column_name')
    key = form_data.get('key')
    value = form_data.get('value')

    # Load the current data
    data = load_data_from_file()

   # Ensure the column exists in the data
    if column_name not in data:
        flash(f"Column '{column_name}' does not exist.", "error")
        return redirect(url_for('admin.admin_window'))

    # Add the key-value pair to the specified column
    if key in data[column_name]:
        flash(f"Key '{key}' already exists in {column_name}.", "error")
    else:
        data[column_name][key] = value
        update_data(data)  # Save the updated data
        flash(f"Key '{key}' with value '{value}' added.", "success")

    return load_admin_page()


@admin.route('/remove_column', methods=['POST'])
@login_required
@admin_login_required
def remove_column():
    form_data = request.form.to_dict()

    # Process the form data (should only contain one key-value pair)
    if len(form_data) != 1:
        flash("Invalid form submission.", "error")
        return redirect(url_for('admin.admin_window'))

    # Extract the column name and value
    column_name, value_to_remove = next(iter(form_data.items()))
    data = load_data_from_file()
    
    # Ensure the column exists in the data
    if column_name in data and value_to_remove in data[column_name]:
        # Remove the value from the column
        del data[column_name][value_to_remove]
        print("data after:", data)
        update_data(data)  # Save the updated data
        flash(f"Value '{value_to_remove}' removed.", "success")
    else:
        flash(f"Value '{value_to_remove}' not found.", "error")
    return load_admin_page()


def load_admin_page():
    # Load data from the file
    data = load_data_from_file()

    # Filter available colors
    available_colors = get_available_colors(data)

    # Render the admin.html template with both `data` and `available_colors`
    return render_template(
        'admin.html',
        data=data,
        available_colors=available_colors
    )

def get_available_colors(data):
    students = data.get('students', {})  # Get the students dictionary
    colors = data.get('colors', {})  # Get the colors dictionary

    # Filter available colors
    available_colors = {
        key: value
        for key, value in colors.items()
        if value not in students.values()  # Check if the color is not assigned to any student
    }

    return available_colors


@admin.route('/excel', methods=['GET'])
@login_required
@admin_login_required
def excel():
    try:
        # Check if the file exists
        if not os.path.exists(EXCEL_FILE):
            raise FileNotFoundError(f"File not found: {EXCEL_FILE}")
        
        # Send the file as a downloadable response
        return send_file(EXCEL_FILE, as_attachment=True)
    except FileNotFoundError as e:
        # Flash an error message if the file is not found
        flash(f"An error occurred while opening the Excel file: {str(e)}", "error")
        return redirect(url_for('admin.admin_window'))
    except Exception as e:
        # Handle other exceptions
        flash(f"An unexpected error occurred: {str(e)}", "error")
        return redirect(url_for('admin.admin_window'))
    
    return load_admin_page()

@admin.route('/slim_excel', methods=['GET'])
@login_required
@admin_login_required
def slim_excel():
    try:
        # Check if the file exists
        if not os.path.exists(ADMIN_EXCEL_FILE):
            raise FileNotFoundError(f"File not found: {ADMIN_EXCEL_FILE}")
        
        # Send the file as a downloadable response
        return send_file(ADMIN_EXCEL_FILE, as_attachment=True)
    except FileNotFoundError as e:
        # Flash an error message if the file is not found
        flash(f"An error occurred while opening the Excel file: {str(e)}", "error")
        return redirect(url_for('admin.admin_window'))
    except Exception as e:
        # Handle other exceptions
        flash(f"An unexpected error occurred: {str(e)}", "error")
        return redirect(url_for('admin.admin_window'))
    
    return load_admin_page()

@admin.route('/clear_excel', methods=['GET', 'POST'])
def clear_excel():
    try:
        # Get the current datetime stamp
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')

        # Extract the directory and file name
        directory, file_name = os.path.split(EXCEL_FILE)
        file_name_without_ext, file_ext = os.path.splitext(file_name)

        # Create the new file name with the timestamp
        new_file_name = f"{file_name_without_ext}_{timestamp}{file_ext}"
        new_file_path = os.path.join(directory, new_file_name)

        # Rename the file
        os.rename(EXCEL_FILE, new_file_path)

        # Flash a success message
        flash(f"Excel file cleared and backed up to {new_file_name}", "success")
    except Exception as e:
        # Flash an error message if something goes wrong
        flash(f"Error renaming Excel file: {str(e)}", "error")

    return load_admin_page()

@admin.route('/reorder_students', methods=['POST'])
def reorder_students():
    from .data_storage import students
    """Reorder a student."""
    reordered_students = request.form.get('reordered_students')
    print("reordered: ", reordered_students)
    print("students: ", students)
    if reordered_students:
        # Parse the reordered_students JSON string into a Python list
        reordered_students = json.loads(reordered_students)
        print("reordered_students: ", reordered_students)

        # Create a new dictionary with the updated order
        updated_students = {student: students[student] for student in reordered_students if student in students}
        print("updated_students: ", updated_students)

        # Update the students dictionary
        students.clear()
        students.update(updated_students)

        # Save the updated dictionary
        save_data()

        flash(f"Students reordered successfully!", "success")
    else:
        flash("Student reorder failed.", "error")

    return load_admin_page()

@admin.route('/reorder_times', methods=['POST'])
def reorder_times():
    from .data_storage import times
    """Reorder a time."""
    reordered_times = request.form.get('reordered_times')
    print("reordered: ", reordered_times)
    print("times: ", times)
    if reordered_times:
        # Parse the reordered_students JSON string into a Python list
        reordered_times = json.loads(reordered_times)
        print("reordered_times: ", reordered_times)
        times.clear()
        times.extend(reordered_times)  # Update the times list with the new order

        
        # Save the updated dictionary
        save_data()
        reorder_excel(times)
        flash(f"Times reordered successfully!", "success")
    else:
        flash("Time reorder failed.", "error")

    return load_admin_page()

import calendar

def reorder_excel(times):
    try:
        # backup excel just in case
        backup_excel()

        # Load the Excel workbook
        wb = openpyxl.load_workbook(EXCEL_FILE)
        admin_wb = openpyxl.load_workbook(ADMIN_EXCEL_FILE)

        # Iterate through all sheets in the workbook
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            admin_sheet = admin_wb[sheet_name]

            # Read the existing data (excluding the first 3 rows)
            data = []
            admin_data = []
            for row in sheet.iter_rows(min_row=5, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column, values_only=True):
                data.append(row)

            for admin_row in admin_sheet.iter_rows(min_row=5, max_row=admin_sheet.max_row, min_col=1, max_col=admin_sheet.max_column, values_only=True):
                admin_data.append(admin_row)

            # Create a mapping of times to rows
            time_to_row = {row[0]: row for row in data}  # Assuming column A contains the times
            admin_time_to_row = {row[0]: row for row in admin_data}  # Assuming column A contains the times

            # Sort the data based on the provided times
            sorted_data = [time_to_row[time] for time in times if time in time_to_row]
            admin_sorted_data = [admin_time_to_row[time] for time in times if time in admin_time_to_row]

            # Clear the rows starting from A4
            for row in sheet.iter_rows(min_row=5, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.value = None
            for admin_row in admin_sheet.iter_rows(min_row=5, max_row=admin_sheet.max_row, min_col=1, max_col=admin_sheet.max_column):
                for admin_cell in admin_row:
                    admin_cell.value = None

            # Write the sorted data back to the sheet starting from A4
            for row_idx, row_data in enumerate(sorted_data, start=5):
                for col_idx, value in enumerate(row_data, start=1):
                    sheet.cell(row=row_idx, column=col_idx, value=value)
            for row_idx, row_data in enumerate(admin_sorted_data, start=5):
                for col_idx, value in enumerate(row_data, start=1):
                    admin_sheet.cell(row=row_idx, column=col_idx, value=value)

            month_name = sheet["A2"].value
            month_number = list(calendar.month_name).index(month_name)

            year = sheet["A3"].value
            sheet = format_worksheet(sheet, year, month_number)
            admin_sheet = format_worksheet(admin_sheet, year, month_number)

        # Save the workbook
        wb.save(EXCEL_FILE)
        admin_wb.save(ADMIN_EXCEL_FILE)
        print(f"Excel file successfully reordered and saved to {EXCEL_FILE}")
    except Exception as e:
        print(f"Error updating {EXCEL_FILE}: {e}")

def backup_excel():
    try:
        # Get the current datetime stamp
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')

        # Extract the directory and file name
        directory, file_name = os.path.split(EXCEL_FILE)
        file_name_without_ext, file_ext = os.path.splitext(file_name)

        # Create the new file name with the timestamp
        new_file_name = f"{file_name_without_ext}_{timestamp}{file_ext}"
        new_file_path = os.path.join(directory, ARCHIVE_FOLDER, new_file_name)

        # Rename the file
        shutil.copy(EXCEL_FILE, new_file_path)


        # ADMIN_EXCEL_FILE
        directory, file_name = os.path.split(ADMIN_EXCEL_FILE)
        file_name_without_ext, file_ext = os.path.splitext(file_name)

        # Create the new file name with the timestamp
        new_file_name = f"{file_name_without_ext}_{timestamp}{file_ext}"
        new_file_path = os.path.join(directory, ARCHIVE_FOLDER, new_file_name)

        # Rename the file
        shutil.copy(ADMIN_EXCEL_FILE, new_file_path)

    except Exception as e:
        # Flash an error message if something goes wrong
        flash(f"Error renaming Excel file: {str(e)}", "error")

