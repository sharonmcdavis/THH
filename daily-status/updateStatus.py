import tkinter as tk
from tkinter import messagebox
import openpyxl
import os
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from tkinter import ttk  # Import ttk for the Separator widget
import json

# File to store the lists
data_file = "app_data.json"

# Function to save the lists to a file
def save_data():
    data = {
        "students": students,
        "times": times,
        "column1": column1,
        "column2": column2,
        "column3": column3,
        "column4": column4,
    }
    with open(data_file, "w") as file:
        json.dump(data, file)

# Function to load the lists from a file
def load_data():
    global students, times, column1, column2, column3, column4
    if os.path.exists(data_file):
        with open(data_file, "r") as file:
            data = json.load(file)
            students = data.get("students", [])
            times = data.get("times", [])
            column1 = data.get("column1", [])
            column2 = data.get("column2", [])
            column3 = data.get("column3", [])
            column4 = data.get("column4", [])
    else:
        # Initialize with default values if the file does not exist
        students = []
        times = []
        column1 = []
        column2 = []
        column3 = []
        column4 = []

# Load data from the file when the app starts
load_data()

# Create the main application window
root = tk.Tk()
root.title("Student Activity Tracker")
root.geometry("1200x800")  # Set the window size

# Variables to track selected buttons
selected_student = tk.StringVar()
selected_time = tk.StringVar()
column1_values = {label: tk.BooleanVar() for label in column1}
column2_values = {label: tk.BooleanVar() for label in column2}
column3_values = {label: tk.BooleanVar() for label in column3}
column4_values = {label: tk.BooleanVar() for label in column4}

# Function to handle toggle button clicks
def toggle_button(var, value):
    if var.get() == value:
        var.set("")  # Deselect if clicked again
    else:
        var.set(value)

# Function to submit the data
def submit_data():
    if not selected_student.get():
        messagebox.showerror("Error", "Please select a student.")
        return
    if not selected_time.get():
        messagebox.showerror("Error", "Please select a time.")
        return

    # Collect selected values
    data = {
        "Student": selected_student.get(),
        "Time": selected_time.get(),
        "Column1": [label for label, var in column1_values.items() if var.get()],
        "Column2": [label for label, var in column2_values.items() if var.get()],
        "Column3": [label for label, var in column3_values.items() if var.get()],
        "Column4": [label for label, var in column4_values.items() if var.get()],
        "Notes": notes_text.get("1.0", "end-1c").strip(),  
    }

    # Write data to Excel
    write_to_excel(data)
    messagebox.showinfo("Success", "Data submitted successfully!")
    reset_buttons()

# Function to write data to an Excel sheet
def write_to_excel(data):
    file_name = "student_activity.xlsx"
    try:
        workbook = openpyxl.load_workbook(file_name)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Activity Log"

    # Write headers if the sheet is empty
    if sheet.max_row == 1:
        sheet.append(["Student", "Time", "Column1", "Column2", "Column3", "Column4", "Notes"])

    # Write the data
    sheet.append([
        data["Student"],
        data["Time"],
        ", ".join(data["Column1"]),
        ", ".join(data["Column2"]),
        ", ".join(data["Column3"]),
        ", ".join(data["Column4"]),
        data["Notes"],  # Add the notes to the Excel file
    ])
    workbook.save(file_name)

# Function to reset all buttons
def reset_buttons():
    selected_student.set("")
    selected_time.set("")
    for var in column1_values.values():
        var.set(False)
    for var in column2_values.values():
        var.set(False)
    for var in column3_values.values():
        var.set(False)
    for var in column4_values.values():
        var.set(False)

# Function to open the Excel file
def open_excel_file():
    file_name = "student_activity.xlsx"
    if os.path.exists(file_name):
        os.startfile(file_name)  # Opens the file with the default application
    else:
        messagebox.showerror("Error", "Excel file not found.")

# Function to convert Excel to PDF and open it
def convert_to_pdf():
    excel_file = "student_activity.xlsx"
    pdf_file = "student_activity.pdf"

    if not os.path.exists(excel_file):
        messagebox.showerror("Error", "Excel file not found.")
        return

    # Create a PDF
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
    pdf = canvas.Canvas(pdf_file, pagesize=letter)
    pdf.setFont("Helvetica", 10)

    # Write data from Excel to PDF
    x, y = 50, 750  # Starting position
    for row in sheet.iter_rows(values_only=True):
        line = " | ".join([str(cell) if cell is not None else "" for cell in row])
        pdf.drawString(x, y, line)
        y -= 20
        if y < 50:  # Create a new page if the content exceeds the page
            pdf.showPage()
            pdf.setFont("Helvetica", 10)
            y = 750

    pdf.save()

    # Open the PDF
    os.startfile(pdf_file)

# Function to open the Admin Functions window
# Add an Exit button to the Admin Window
def open_admin_window():
    """Open the admin window for managing data."""
    admin_window = tk.Toplevel(root)
    admin_window.title("Admin Functions")
    admin_window.geometry("600x500")

    admin_window.transient(root)
    admin_window.grab_set()
    admin_window.focus_set()

    # Add widgets for managing students
    tk.Label(admin_window, text="Manage Students", font=("Arial", 12, "bold")).grid(row=0, column=0, sticky="w", pady=5)
    student_listbox = tk.Listbox(admin_window, height=6, width=30)
    student_listbox.grid(row=1, column=0, sticky="w", pady=5)
    update_listbox(student_listbox, students)  # Populate the listbox with existing students
    student_entry = tk.Entry(admin_window, width=30)
    student_entry.grid(row=2, column=0, sticky="w", pady=5)

    # Add Student Button with validation
    tk.Button(
        admin_window,
        text="Add Student",
        command=lambda: add_student(student_entry, students, student_listbox),
        bg="green",
        fg="white"
    ).grid(row=2, column=1, sticky="w", padx=5)

    # Remove Student Button with validation
    tk.Button(
        admin_window,
        text="Remove Student",
        command=lambda: remove_student(student_listbox, students),
        bg="red",
        fg="white"
    ).grid(row=2, column=2, sticky="w", padx=5)

    # Exit Button (Save and Close)
    tk.Button(
        admin_window,
        text="Exit",
        command=lambda: save_and_close(admin_window),
        bg="light gray",
        fg="black"
    ).grid(row=0, column=2, sticky="ne", padx=10, pady=10)

    # When the admin window is closed, refresh the main window
    def on_admin_close():
        refresh_main_window()
        admin_window.destroy()

    admin_window.protocol("WM_DELETE_WINDOW", on_admin_close)  # Handle the close event


def add_student(entry_widget, target_list, listbox):
    """Add a student to the list with validation."""
    item = entry_widget.get().strip()
    if not item:
        messagebox.showerror("Error", "Please enter a student name.")
        return
    if item in target_list:
        messagebox.showerror("Error", f"'{item}' already exists!")
        return
    target_list.append(item)
    update_listbox(listbox, target_list)  # Update the listbox
    save_data()  # Save the updated data
    messagebox.showinfo("Success", f"'{item}' added successfully!")
    entry_widget.delete(0, tk.END)


def remove_student(listbox, target_list):
    """Remove a student from the list with validation."""
    selected_index = listbox.curselection()
    if not selected_index:
        messagebox.showerror("Error", "Please select a student to remove.")
        return
    item = listbox.get(selected_index)
    target_list.remove(item)
    update_listbox(listbox, target_list)  # Update the listbox
    save_data()  # Save the updated data
    messagebox.showinfo("Success", f"'{item}' removed successfully!")

# Function to refresh the main window's data
def refresh_main_window():
    """Refresh the main window with updated data."""
    # Clear all widgets in the main window
    for widget in table_frame.winfo_children():
        widget.destroy()

    # Recreate the main window layout
    create_main_window_table()

def save_and_close(admin_window):
    """Save data and close the admin window."""
    save_data()  # Save the updated data to the JSON file
    refresh_main_window()  # Refresh the main window with updated data
    admin_window.destroy()

# Add a callback to refresh the main window when the admin window is closed
def open_admin_window():
    admin_window = tk.Toplevel(root)
    admin_window.title("Admin Functions")
    admin_window.geometry("600x500")

    admin_window.transient(root)
    admin_window.grab_set()
    admin_window.focus_set()

    admin_window.focus_set()

    # Add widgets for managing students
    tk.Label(admin_window, text="Manage Students", font=("Arial", 12, "bold")).grid(row=0, column=0, sticky="w", pady=5, padx=20)  # Add left padding
    student_listbox = tk.Listbox(admin_window, height=6, width=30)
    student_listbox.grid(row=1, column=0, sticky="w", pady=(5, 20), padx=20)  # Add extra space below the listbox and left padding
    update_listbox(student_listbox, students)  # Populate the listbox with existing students

    student_entry = tk.Entry(admin_window, width=30)
    student_entry.grid(row=2, column=0, sticky="w", pady=(10, 5), padx=20)  # Add extra space above the text box and left padding

    # Add Student Button
    tk.Button(
        admin_window,
        text="Add Student",
        command=lambda: add_item(student_entry, students, student_listbox),
        bg="green",
        fg="white",
        width=15,  # Set consistent width
        height=2   # Set consistent height
    ).grid(row=2, column=1, sticky="w", padx=15)

    # Remove Student Button (aligned to the top of the selection box)
    tk.Button(
        admin_window,
        text="Remove Student",
        command=lambda: remove_item(student_entry, students, student_listbox),
        bg="purple",
        fg="white",
        width=15,  # Set consistent width
        height=2   # Set consistent height
    ).grid(row=1, column=1, sticky="s", pady=15, padx=15)  # Align to the top of the listbox

    # Exit Button
    tk.Button(
        admin_window,
        text="Exit",
        command=lambda: save_and_close(admin_window),
        bg="gray",
        fg="white",
        width=15,  # Set consistent width
        height=2   # Set consistent height
    ).grid(row=0, column=5, pady=20, sticky="ne", padx=10)

    # When the admin window is closed, refresh the main window
    def on_admin_close():
        refresh_main_window()
        admin_window.destroy()

    admin_window.protocol("WM_DELETE_WINDOW", on_admin_close)  # Handle the close event

# Function to create the main window's table
def create_main_window_table():
    # Create the table dynamically based on the updated data
    tk.Label(table_frame, text="Students", font=("Arial", 12, "bold")).grid(row=0, column=0, padx=20, pady=5)  # Add left padding
    for i, student in enumerate(students):
        tk.Button(table_frame, text=student, command=lambda s=student: print(f"Selected student: {s}")).grid(row=i + 1, column=0, padx=20, pady=5)  # Add left padding

# Function to create the main window's table
def create_main_window_table():
    # Create the table dynamically based on the updated data
    tk.Label(table_frame, text="Students", font=("Arial", 12, "bold")).grid(row=0, column=0, padx=5, pady=5)
    for i, student in enumerate(students):
        tk.Button(table_frame, text=student, command=lambda s=student: print(f"Selected student: {s}")).grid(row=i + 1, column=0, padx=5, pady=5)

    tk.Label(table_frame, text="Times", font=("Arial", 12, "bold")).grid(row=0, column=1, padx=5, pady=5)
    for j, time in enumerate(times):
        tk.Button(table_frame, text=time, command=lambda t=time: print(f"Selected time: {t}")).grid(row=j + 1, column=1, padx=5, pady=5)

    # Add column options dynamically
    for col_index, column in enumerate([column1, column2, column3, column4], start=2):
        tk.Label(table_frame, text=f"Column {col_index - 1}", font=("Arial", 12, "bold")).grid(row=0, column=col_index, padx=5, pady=5)
        for i, option in enumerate(column):
            tk.Checkbutton(table_frame, text=option).grid(row=i + 1, column=col_index, padx=5, pady=5)

# Helper function to update the listbox with current items
def update_listbox(listbox, items):
    listbox.delete(0, tk.END)  # Clear the listbox
    for item in items:
        listbox.insert(tk.END, item)  # Add each item to the listbox

# Helper function to add an item to a list and update the listbox
def add_item(entry_widget, target_list, listbox):
    item = entry_widget.get().strip()
    if item and item not in target_list:
        target_list.append(item)
        update_listbox(listbox, target_list)  # Update the listbox
        save_data()  # Save the updated data
        messagebox.showinfo("Success", f"'{item}' added successfully!")
        entry_widget.delete(0, tk.END)
    elif item in target_list:
        messagebox.showerror("Error", f"'{item}' already exists!")
    else:
        messagebox.showerror("Error", "Please enter a valid item.")

# Helper function to modify an item in a list and update the listbox
def modify_item(entry_widget, target_list, listbox):
    selected_index = listbox.curselection()
    if selected_index:
        old_item = listbox.get(selected_index)
        new_item = entry_widget.get().strip()
        if new_item and new_item not in target_list:
            target_list[target_list.index(old_item)] = new_item
            update_listbox(listbox, target_list)  # Update the listbox
            save_data()  # Save the updated data
            messagebox.showinfo("Success", f"'{old_item}' modified to '{new_item}' successfully!")
            entry_widget.delete(0, tk.END)
        elif new_item in target_list:
            messagebox.showerror("Error", f"'{new_item}' already exists!")
        else:
            messagebox.showerror("Error", "Please enter a valid item.")
    else:
        messagebox.showerror("Error", "Please select an item to modify.")

# Helper function to remove an item from a list and update the listbox
def remove_item(entry_widget, target_list, listbox):
    selected_index = listbox.curselection()
    if selected_index:
        item = listbox.get(selected_index)
        target_list.remove(item)
        update_listbox(listbox, target_list)  # Update the listbox
        save_data()  # Save the updated data
        messagebox.showinfo("Success", f"'{item}' removed successfully!")
        entry_widget.delete(0, tk.END)
    else:
        messagebox.showerror("Error", "Please select an item to remove.")

# Helper function to get the appropriate column list
def get_column_list(column_name):
    if column_name == "Column1":
        return column1
    elif column_name == "Column2":
        return column2
    elif column_name == "Column3":
        return column3
    elif column_name == "Column4":
        return column4
        
# --------------------------------------------------------------------------------

# Vertical Students (aligned to the left)
student_frame = tk.Frame(root)
student_frame.pack(side=tk.LEFT, padx=10, pady=10, anchor="n")  # Align to the left
tk.Label(student_frame, text="Students").pack(pady=5)
for student in students:
    btn = tk.Radiobutton(
        student_frame, text=student, variable=selected_student, value=student,
        indicatoron=False, width=15, height=2, bg="lightblue", selectcolor="blue"
    )
    btn.pack(pady=2)

# Vertical Times (aligned to the left)
time_frame = tk.Frame(root)
time_frame.pack(side=tk.LEFT, padx=10, pady=10, anchor="n")  # Align to the left
tk.Label(time_frame, text="Times").pack(pady=5)
for time in times:
    btn = tk.Radiobutton(
        time_frame, text=time, variable=selected_time, value=time,
        indicatoron=False, width=15, height=2, bg="lightgreen", selectcolor="green"
    )
    btn.pack(pady=2)

# Create the table with toggle buttons
table_frame = tk.Frame(root)
table_frame.pack(pady=25, anchor="w")  # Align the entire table to the left

# Column 1
col1_frame = tk.Frame(table_frame)
col1_frame.grid(row=0, column=0, padx=10, sticky="w")  # Position next to the "time" column
for label in column1:
    btn = tk.Checkbutton(
        col1_frame, text=label, variable=column1_values[label],
        width=30, height=1, anchor="w"  # Align text and checkbox to the left
    )
    btn.pack(anchor="w", pady=2)

# Add a horizontal line (separator) after Column 1 options
separator = ttk.Separator(table_frame, orient="horizontal")
separator.grid(row=1, column=0, sticky="ew", pady=10)  # Stretch the line horizontally

# Column 2
col2_frame = tk.Frame(table_frame)
col2_frame.grid(row=2, column=0, padx=10, sticky="w")  # Position next to Column 1
for label in column2:
    btn = tk.Checkbutton(
        col2_frame, text=label, variable=column2_values[label],
        width=30, height=1, anchor="w"  # Align text and checkbox to the left
    )
    btn.pack(anchor="w", pady=2)

# Add a horizontal line (separator) after Column 1 options
separator = ttk.Separator(table_frame, orient="horizontal")
separator.grid(row=3, column=0, sticky="ew", pady=10)  # Stretch the line horizontally

# Column 3
col3_frame = tk.Frame(table_frame)
col3_frame.grid(row=4, column=0, padx=10, sticky="w")  # Position next to Column 2
for label in column3:
    btn = tk.Checkbutton(
        col3_frame, text=label, variable=column3_values[label],
        width=30, height=1, anchor="w"  # Align text and checkbox to the left
    )
    btn.pack(anchor="w", pady=2)

# Add a horizontal line (separator) after Column 1 options
separator = ttk.Separator(table_frame, orient="horizontal")
separator.grid(row=5, column=0, sticky="ew", pady=10)  # Stretch the line horizontally

# Column 4
col4_frame = tk.Frame(table_frame)
col4_frame.grid(row=6, column=0, padx=10, sticky="w")  # Position next to Column 3
for label in column4:
    btn = tk.Checkbutton(
        col4_frame, text=label, variable=column4_values[label],
        width=30, height=1, anchor="w"  # Align text and checkbox to the left
    )
    btn.pack(anchor="w", pady=2)


# Free-form Notes Section
notes_frame = tk.Frame(table_frame)
notes_frame.grid(row=6, column=0, columnspan=4, pady=20, sticky="w")  # Align below other options

tk.Label(notes_frame, text="Notes:", font=("Arial", 12)).grid(row=0, column=0, sticky="w")
notes_text = tk.Text(notes_frame, height=5, width=50, wrap="word", font=("Arial", 10))
notes_text.grid(row=1, column=0, pady=5, sticky="w")


# Submit, View Excel, and View PDF buttons (right column)
button_frame = tk.Frame(table_frame)
button_frame.grid(row=0, column=5, rowspan=10, padx=20, pady=10, sticky="n")  # Align to the top-right

submit_button = tk.Button(button_frame, text="Submit", command=submit_data, bg="purple", fg="white", height=2, width=20)
submit_button.grid(row=0, column=0, pady=5)  # Add vertical spacing between buttons

# Add Exit Button to Main Window
exit_button = tk.Button(button_frame, text="Exit Application", command=root.quit, bg="green", fg="white", height=2, width=20)
exit_button.grid(row=1, column=0, pady=5)  # Add vertical spacing between buttons

# Add a button on the main screen to access Admin Functions
admin_button = tk.Button(button_frame, text="Admin Functions", command=open_admin_window, bg="gray", fg="white", height=2, width=20)
admin_button.grid(row=2, column=0, pady=5) 

# view_excel_button = tk.Button(button_frame, text="View Excel", command=open_excel_file, bg="blue", fg="white", height=2, width=20)
# view_excel_button.grid(row=1, column=0, pady=5)

# view_pdf_button = tk.Button(button_frame, text="View PDF", command=convert_to_pdf, bg="green", fg="white", height=2, width=20)
# view_pdf_button.grid(row=2, column=0, pady=5)

# Run the application
root.mainloop()