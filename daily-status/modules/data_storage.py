# data_storage.py
import json
from tkinter import messagebox
import os
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from data_loader import load_data_from_file, save_data_to_file

# Global variables
DATA_FILE = "app_data.json" 
EXCEL_FILE = "student_activity.xlsx"
PDF_FILE = "student_activity.pdf"
students = []
times = []
column1 = []
column2 = []
column3 = []
column4 = []

def initialize_data():
    """Load data from the JSON file into global variables."""
    global students, times, column1, column2, column3, column4
    try:
        with open("app_data.json", "r") as file:
            data = json.load(file)
            students = data.get("students", [])
            times = data.get("times", [])
            column1 = data.get("column1", {})  # Load as a dictionary
            column2 = data.get("column2", {})  # Load as a dictionary
            column3 = data.get("column3", {})  # Load as a dictionary
            column4 = data.get("column4", {})  # Load as a dictionary
    except FileNotFoundError:
        print("app_data.json not found. Using default values.")
        students = []
        times = []
        column1 = {}
        column2 = {}
        column3 = {}
        column4 = []

def save_data():
    data = {
        "students": students,
        "times": times,
        "column1": column1,
        "column2": column2,
        "column3": column3,
        "column4": column4,
    }
    print("")
    print("save_data:")
    print(data)
    try:
        with open(DATA_FILE, "w") as json_file:
            json.dump(data, json_file, indent=4)
        print(f"Data successfully saved to {DATA_FILE}")
    except Exception as e:
        print(f"Error saving data to {DATA_FILE}: {e}")

def add_item(entry_widget, data_list, listbox_widget, save_callback, global_var_name=None):
    """Add an item to the data list, update the listbox, and save to the global variable."""
    new_item = entry_widget.get().strip()  # Get the text from the entry widget
    if new_item and new_item not in data_list:  # Check if the item is not empty and not already in the list
        data_list.append(new_item)  # Add the new item to the data list
        update_listbox(listbox_widget, data_list)  # Update the listbox with the new data

        # Update the global variable if a global_var_name is provided
        if global_var_name:
            globals()[global_var_name] = data_list

        save_callback()  # Save the updated data to the JSON file
        entry_widget.delete(0, "end")  # Clear the entry widget after adding the item
    else:
        # Optionally, show a message if the item is empty or already exists
        print("Item is either empty or already exists in the list.")

def remove_item(listbox_widget, data_list, save_callback, global_var_name=None):
    """Remove the selected item from the data list, update the listbox, and save to the global variable."""
    selected_index = listbox_widget.curselection()
    if selected_index:
        item_to_remove = listbox_widget.get(selected_index)
        data_list.remove(item_to_remove)  # Remove the item from the data list
        update_listbox(listbox_widget, data_list)  # Update the listbox with the new data

        # Update the global variable if a global_var_name is provided
        if global_var_name:
            globals()[global_var_name] = data_list

        save_callback()  # Save the updated data to the JSON file
    else:
        # Optionally, show a message if no item is selected
        print("No item selected for removal.")
        
def update_listbox(listbox_widget, data_list):
    """Update the listbox with the latest data."""
    listbox_widget.delete(0, "end")  # Clear the listbox
    for item in data_list:
        listbox_widget.insert("end", item)  # Add updated items

# Function to write data to an Excel sheet
import openpyxl
from openpyxl.styles import Alignment
from datetime import datetime
import os
from openpyxl import load_workbook
from fpdf import FPDF
from openpyxl.styles import PatternFill

def apply_alternate_shading(sheet):
    """Apply alternate shading to every 5 columns in the Excel sheet."""
    # Define fill colors for shading
    fill_color_1 = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")  # Light gray
    fill_color_2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White (no shading)

    # Iterate through columns in groups of 5
    for col_idx in range(2, sheet.max_column + 1):
        # Determine the shading group (alternate every 5 columns)
        shading_group = (col_idx - 1) // 4
        fill_color = fill_color_1 if shading_group % 2 == 0 else fill_color_2

        # Apply the fill color to all rows in the column
        for row_idx in range(1, sheet.max_row + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.fill = fill_color

def write_to_excel(self, data):
    """Write data to an Excel file with student-specific sheets for each month."""
    # Get the current date
    today = datetime.now()
    current_month = today.strftime("%B")  # Full month name (e.g., "December")
    current_day = today.day  # Day of the month (e.g., 31)

    # Load the workbook if it exists, otherwise create a new one
    if os.path.exists(EXCEL_FILE):
        workbook = openpyxl.load_workbook(EXCEL_FILE)
    else:
        workbook = openpyxl.Workbook()
        # Remove the default sheet created by openpyxl
        default_sheet = workbook.active
        workbook.remove(default_sheet)

    # Get the student's name
    student_name = data["Student"]

    # Create the sheet name for the current student and month
    sheet_name = f"{student_name}-{current_month}"

    # Check if the sheet already exists, otherwise create it
    if sheet_name not in workbook.sheetnames:
        sheet = workbook.create_sheet(title=sheet_name)

        # Add the student's name in the first row
        sheet["A1"] = student_name
        sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

        # Add headers in the second row
        sheet["A2"] = "Dates"
        for day in range(1, 32):  # Maximum days in a month
            col_letter = openpyxl.utils.get_column_letter(day + 1)  # Start from column B
            sheet[f"{col_letter}2"] = str(day)

        # Add times in column A starting from row 3
        for idx, time in enumerate(self.times, start=3):
            sheet[f"A{idx}"] = time

        # Apply alternate shading
        apply_alternate_shading(sheet)

    else:
        # If the sheet already exists, load it
        sheet = workbook[sheet_name]

    # Find the column for the current day
    day_column = None
    for col in range(2, sheet.max_column + 1):  # Start from column B
        col_letter = openpyxl.utils.get_column_letter(col)
        if sheet[f"{col_letter}2"].value == str(current_day):
            day_column = col_letter
            break

    if not day_column:
        raise ValueError(f"Could not find the column for day {current_day} in sheet {sheet_name}.")

    # Find the row for the selected time
    time_row = None
    for row in range(3, sheet.max_row + 1):  # Start from row 3
        if sheet[f"A{row}"].value == data["Time"]:
            time_row = row
            break

    if not time_row:
        raise ValueError(f"Could not find the row for time {data['Time']} in sheet {sheet_name}.")

    # Filter out "UNSELECTED" values and concatenate the remaining column values
    columns_to_concatenate = [
        data[key].strip() for key in ["Column 1", "Column 2", "Column 3", "Column 4"]
        if data[key] != "UNSELECTED" and data[key].strip()  # Exclude empty or whitespace-only values
    ]
    concatenated_values = "\n".join(columns_to_concatenate)

    # Add notes if they exist
    if data["Notes"]:
        notes = data["Notes"].strip()  # Remove leading/trailing whitespace from notes
        if notes:  # Only add notes if they are not empty
            concatenated_values += f"\n{notes}"

    print(day_column)
    print(time_row)
    print(concatenated_values)

    # Write the concatenated values to the appropriate cell
    sheet[f"{day_column}{time_row}"] = concatenated_values

    # Save the workbook
    workbook.save(EXCEL_FILE)
    enable_text_wrapping(EXCEL_FILE)
    return True

from openpyxl import load_workbook
from openpyxl.styles import Alignment

def enable_text_wrapping(excel_file):
    # Load the workbook
    workbook = load_workbook(excel_file)

    # Iterate through all sheets in the workbook
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Iterate through all cells in the sheet
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and "\n" in str(cell.value):  # Check if the cell contains a newline
                    cell.alignment = Alignment(wrap_text=True)  # Enable text wrapping

    # Save the workbook
    workbook.save(excel_file)
    print(f"Text wrapping enabled in {excel_file}")

# Function to open the Excel file
def open_excel_file():
    print("excel_file()")
    if os.path.exists(EXCEL_FILE):
        os.startfile(EXCEL_FILE)  # Opens the file with the default application
    else:
        messagebox.showerror("Error", "Excel file not found.")

# def convert_to_pdf():
#     print("convert_to_pdf")
#     # Load the workbook
#     workbook = load_workbook(EXCEL_FILE)

#     # Create a PDF object
#     pdf = PDF()
#     pdf.set_auto_page_break(auto=True, margin=15)

#     # Iterate through all sheets in the workbook
#     for sheet_name in workbook.sheetnames:
#         sheet = workbook[sheet_name]

#         # Add a new page for each sheet
#         pdf.add_page()
#         pdf.set_font("Arial", "B", 12)
#         pdf.cell(0, 10, f"Sheet: {sheet_name}", ln=True, align="C")
#         pdf.ln(5)

#         # Extract data from the sheet
#         data = list(sheet.iter_rows(values_only=True))

#         # Determine column widths (adjust as needed)
#         col_widths = [20] + [15] * (len(data[0]) - 1) if data else [20]

#         # Add table to the PDF
#         pdf.add_table(data, col_widths)

#     # Save the PDF
#     pdf.output(PDF_FILE)
#     print(f"PDF saved to {PDF_FILE}")
#     os.startfile(PDF_FILE)

# class PDF(FPDF):
#     def header(self):
#         self.set_font("Arial", "B", 12)
#         self.cell(0, 10, "Excel to PDF Report", border=False, ln=True, align="C")
#         self.ln(5)

#     def add_table(self, data, col_widths, col_headers=None):
#         self.set_font("Arial", size=10)

#         # Add column headers if provided
#         if col_headers:
#             for i, header in enumerate(col_headers):
#                 self.cell(col_widths[i], 10, header, border=1, align="C")
#             self.ln()

#         # Add rows of data
#         for row in data:
#             for i, cell in enumerate(row):
#                 self.cell(col_widths[i], 10, str(cell) if cell is not None else "", border=1, align="C")
#             self.ln()

from openpyxl import load_workbook
from fpdf import FPDF


class PDF(FPDF):
    def __init__(self, orientation="L", unit="mm", format="A4"):
        super().__init__(orientation, unit, format)
        self.set_auto_page_break(auto=True, margin=10)

    # def header(self):
    #     self.set_font("Arial", "B", 12)
    #     self.cell(0, 10, "Excel to PDF Report", border=False, ln=True, align="C")
    #     self.ln(5)

    def add_table(self, data, page_width, col_headers=None):
        self.set_font("Arial", size=8)

        # Calculate column widths dynamically
        num_columns = len(data[0]) if data else 1
        col_width = page_width / num_columns

        # Add column headers if provided
        if col_headers:
            for header in col_headers:
                self.cell(col_width, 10, header, border=1, align="C")
            self.ln()

        # Add rows of data
        for row in data:
            for cell in row:
                self.cell(col_width, 10, str(cell) if cell is not None else "", border=1, align="C")
            self.ln()


def convert_to_pdf():
    # Load the workbook
    workbook = load_workbook(EXCEL_FILE)

    # Create a PDF object in landscape mode
    pdf = PDF(orientation="L")
    pdf.set_auto_page_break(auto=True, margin=10)

    # Get the page width (usable width for content)
    page_width = pdf.w - 20  # Subtracting margins (10mm on each side)

    # Iterate through all sheets in the workbook
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Add a new page for each sheet
        pdf.add_page()
        pdf.set_font("Arial", "B", 10)
        pdf.cell(0, 10, f"{sheet_name}", ln=True, align="C")
        pdf.ln(5)

        # Extract data from the sheet
        data = list(sheet.iter_rows(values_only=True))

        # Add table to the PDF
        pdf.add_table(data, page_width)

    # Save the PDF
    pdf.output(PDF_FILE)
    print(f"PDF saved to {PDF_FILE}")
    os.startfile(PDF_FILE)
